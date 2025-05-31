import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from services.encryption import decrypt_data
from pathlib import Path

def detect_file_type(file_path: Path) -> str:
    """
    Determine the type of a file based on its extension.

    This function checks the suffix of the given Path object and returns a
    string indicating whether it is a PDF or Excel file. If the extension is
    not recognized, it raises a ValueError.

    Parameters
    ----------
    file_path : Path
        A pathlib Path pointing to the file whose type needs detection.

    Returns
    -------
    str
        "pdf" if the file has a .pdf extension (case-insensitive),
        "excel" if the file has a .xls or .xlsx extension.

    Raises
    ------
    ValueError
        If the file extension is not one of the supported types.
    """
    if file_path.suffix.lower() == ".pdf":
        return "pdf"
    elif file_path.suffix.lower() in [".xls", ".xlsx"]:
        return "excel"
    else:
        raise ValueError("Unsupported file type")


def parse_pdf(file_path: Path) -> list[dict]:
    """
    Read and decrypt a PDF file, then extract text from each page.

    This function opens the encrypted PDF at the given path, decrypts its
    contents, and uses PyMuPDF (fitz) to extract text from every page.
    It returns a list of dictionaries, each representing one page with its
    text and associated metadata.

    Parameters
    ----------
    file_path : Path
        A pathlib Path pointing to the encrypted PDF file to parse.

    Returns
    -------
    list[dict]
        A list where each element is a dictionary containing:
            - "type":       The string "pdf" to indicate file type.
            - "page":       The 1-based page number (int).
            - "text":       The extracted text from that page (str).
            - "source_file": The original filename (str).
    """
    results = []
    with open(file_path, "rb") as f:
        encrypted_bytes = f.read()
    decrypted = decrypt_data(encrypted_bytes)

    # Open the decrypted PDF from memory
    doc = fitz.open(stream=decrypted, filetype="pdf")
    for i, page in enumerate(doc):
        text = page.get_text()
        results.append({
            "type":        "pdf",
            "page":        i + 1,
            "text":        text,
            "source_file": file_path.name
        })
    return results


def parse_excel(file_path: Path) -> list[dict]:
    """
    Read and decrypt an Excel file, then extract text from each sheet.

    This function decrypts the Excel file at the given path, writes the
    decrypted bytes to a temporary .xlsx file, reads each sheet into a pandas
    DataFrame, and converts it to a string. It returns a list of dictionaries
    with sheet-level text and metadata. The temporary file is deleted after use.

    Parameters
    ----------
    file_path : Path
        A pathlib Path pointing to the encrypted Excel (.xls or .xlsx) file.

    Returns
    -------
    list[dict]
        A list where each element is a dictionary containing:
            - "type":       The string "excel" to indicate file type.
            - "sheet":      The sheet name (str).
            - "text":       The DataFrame contents as a string (str).
            - "source_file": The original filename (str).
    """
    results = []
    with open(file_path, "rb") as f:
        encrypted_bytes = f.read()
    decrypted = decrypt_data(encrypted_bytes)

    # Write decrypted bytes to a temporary file for pandas to read
    tmp_path = file_path.with_suffix(".decrypted.xlsx")
    with open(tmp_path, "wb") as tmp:
        tmp.write(decrypted)

    wb = load_workbook(filename=tmp_path, read_only=True)
    for sheet in wb.sheetnames:
        df = pd.read_excel(tmp_path, sheet_name=sheet)
        results.append({
            "type":        "excel",
            "sheet":       sheet,
            "text":        df.to_string(index=False),
            "source_file": file_path.name
        })

    tmp_path.unlink()  # Delete the temporary file
    return results
